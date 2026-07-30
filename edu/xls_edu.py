import os
import logging
import chardet
import openpyxl
try:
    import xlrd
except Exception:
    xlrd = None
import pandas as pd
import csv
from db.db_operations import insert_data, delete_data
try:
    from utils.llm_agent import clean_data_with_llm
except ImportError:
    clean_data_with_llm = None
from backend.config import Config
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
from utils.logging_util import LoggerSingleton
from db.db_job_managers import AsyncJobManager, AsyncJobProgress
import asyncio
from edu.file_text_extract_timeout import await_file_text_extract
import time
from typing import Any, Dict, List

# 로거 설정
logger = LoggerSingleton.get_logger(logger_name="edu.xls", level=logging.INFO)

embedding_model = OpenAIEmbeddings(openai_api_key=Config.OPENAI_API_KEY)
FILE_EMBEDDING_BATCH_SIZE = max(1, int(getattr(Config, "FILE_EMBEDDING_BATCH_SIZE", 5) or 5))


def _coerce_positive_int_text(value: Any, fallback: int = 1) -> str:
    """Return a positive integer string for PG chunk/page number columns."""
    try:
        if isinstance(value, bool):
            raise ValueError("bool is not a valid chunk number")
        if isinstance(value, int):
            parsed = value
        elif isinstance(value, float):
            if not value.is_integer():
                raise ValueError("non-integer float")
            parsed = int(value)
        else:
            text = str(value or "").strip()
            if not text:
                raise ValueError("blank value")
            if text.endswith(".0") and text[:-2].isdigit():
                text = text[:-2]
            if not text.isdigit():
                raise ValueError(f"non-numeric value: {text!r}")
            parsed = int(text)
        if parsed <= 0:
            raise ValueError("non-positive value")
    except Exception:
        parsed = max(1, int(fallback or 1))
    return str(parsed)


def _normalize_tabular_row(row: List[Any]) -> List[str]:
    """빈 셀/공백 셀을 정리하고 문자열 리스트로 표준화한다."""
    out: List[str] = []
    for cell in row or []:
        try:
            text = str(cell).strip()
        except Exception:
            text = ""
        if text:
            out.append(text)
    return out


def _first_meaningful_row(rows: List[List[Any]]) -> List[str]:
    """CSV 구분자 추론용: 비어 있지 않은 첫 행을 안전하게 반환한다."""
    for row in rows or []:
        normalized = _normalize_tabular_row(list(row or []))
        if normalized:
            return normalized
    return []


def _load_excel_sheets_dict_pandas_fallback(file_path: str) -> Dict[str, List[List[Any]]]:
    """Fallback loader for xls/xlsx/csv when the primary reader fails."""
    file_extension = os.path.splitext(file_path)[1].lower()
    sheets_data: Dict[str, List[List[Any]]] = {}

    if file_extension in (".xlsx", ".xls"):
        raw_sheets = pd.read_excel(file_path, sheet_name=None, header=None, dtype=str)
        for sheet_name, df in (raw_sheets or {}).items():
            if df is None or df.empty:
                continue
            sheet_rows: List[List[Any]] = []
            for _, row in df.fillna("").iterrows():
                row_data = _normalize_tabular_row(list(row.tolist()))
                if row_data:
                    sheet_rows.append(row_data)
            if sheet_rows:
                sheets_data[str(sheet_name)] = sheet_rows
        return sheets_data

    if file_extension == ".csv":
        df = pd.read_csv(
            file_path,
            header=None,
            dtype=str,
            encoding_errors="ignore",
            engine="python",
            on_bad_lines="skip",
            sep=None,
        )
        csv_rows: List[List[Any]] = []
        for _, row in df.fillna("").iterrows():
            row_data = _normalize_tabular_row(list(row.tolist()))
            if row_data:
                csv_rows.append(row_data)
        if csv_rows:
            sheets_data["csv"] = csv_rows
        return sheets_data

    raise ValueError(f"Unsupported extension for fallback: {file_extension}")


# ✅ 개별 청크를 비동기로 처리하는 함수 (url_edu.py 방식 적용)
async def process_single_chunk_async(
    chunk_info: dict,
    content: str,
    subject: str,
    sheet_name: str,
    table_name: str,
    dbname: str,
    memo: str,
    job_id: str,
    job_manager: AsyncJobManager,
    personal_info_filter: str = "N",  # 개인정보 필터링 옵션 추가
    content_type: str = "file",
):
    """개별 청크를 비동기로 처리 (임베딩 + 저장)"""
    try:
        logger.info(f"= = = = = = = = = = = = = = = = [XLS 청크 처리] content: {content}, memo: '{memo}' (type: {type(memo)})")
        # ✅ 청크 처리 전 취소 상태 확인
        status = await job_manager.get_job_status(job_id)
        if status == "cancel":
            logger.info(f"엑셀 개별 청크 처리 중 작업 취소됨: job_id={job_id}, chunk={chunk_info['chunk_idx']}")
            return {"status": "cancelled", "chunk_idx": chunk_info['chunk_idx']}
        
        # 청크 생성
        chunk_text = create_chunk_from_rows(
            sheet_name=sheet_name,
            header_rows=chunk_info["header_rows"], 
            chunk_rows=chunk_info["data_rows"],
            chunk_idx=chunk_info["chunk_idx"]
        )
        if personal_info_filter == "Y":
            logger.info(f"XLS 개별 청크 처리 중 개인정보 필터링 적용: job_id={job_id}, chunk={chunk_info['chunk_idx']}")
            from utils.dlp_api import check_pii_content
            original_chunk_text = chunk_text  # 원본 청크 저장
            pii_result = check_pii_content(chunk_text)
            
            # 마스킹된 텍스트를 chunk_text 변수에 할당
            if pii_result["success"]:
                chunk_text = pii_result["masked_text"]
                

            else:
                # PII 검사 실패 시 에러 로깅
                logger.error(f"PII 검사 실패: {pii_result.get('error', 'Unknown error')}")
                # 실패 시 원본 텍스트 유지
        
        # ✅ td_*_training_data의 chunk_num/page_num 컬럼은 "숫자"로 저장한다.
        # - chunk_num: 파일 내 전역 청크 순번(1..N) (정렬/집계 용이)
        # - page_num: 시트 순번(1..S)
        fallback_chunk_num = chunk_info.get("chunk_idx") or 1
        chunk_num = _coerce_positive_int_text(chunk_info.get("chunk_num"), fallback=fallback_chunk_num)
        # td_*_training_data의 page_num 컬럼이 TEXT인 환경(asyncpg)과 호환되도록 문자열로 저장한다.
        # (숫자 의미는 유지: "1", "2", ...)
        page_num = _coerce_positive_int_text(chunk_info.get("sheet_idx"), fallback=1)
        chunk_with_metadata = chunk_text
        
        # ✅ 비동기 임베딩 생성
        embedding = await embedding_model.aembed_query(chunk_with_metadata)
        embedding_array = f"[{','.join(map(str, embedding))}]"

        # ✅ DB 저장 전 한 번 더 취소 상태 확인
        status = await job_manager.get_job_status(job_id)
        if status == "cancel":
            logger.info(f"엑셀 DB 저장 전 작업 취소됨: job_id={job_id}, chunk={chunk_info['chunk_idx']}")
            return {"status": "cancelled", "chunk_idx": chunk_info['chunk_idx']}

        # ✅ 비동기 DB 저장
        logger.info(f"= = = = = = = = = = = = = = [XLS DB 저장] content: {content}, chunk_num: {chunk_num}, memo: '{memo}'")
        await insert_data(
            table=table_name,
            data={
                "content": content,
                "subject": subject,
                "chunk_num": chunk_num,
                "page_num": page_num,
                "memo": memo,
                "content_type": content_type,
                "text_data": chunk_with_metadata,
                "embedding": embedding_array,
            },
            dbname=dbname,
        )

        return {
            "status": "success", 
            "chunk_idx": chunk_info['chunk_idx'],
            "rows_in_chunk": len(chunk_info["data_rows"]),
            "sheet_name": sheet_name
        }

    except Exception as e:
        logger.error(f"엑셀 청크 처리 오류 ({content}, sheet: {sheet_name}, chunk {chunk_info['chunk_idx']}): {e}")
        return {
            "status": "error", 
            "chunk_idx": chunk_info['chunk_idx'], 
            "error": str(e),
            "sheet_name": sheet_name
        }


# ✅ 청크들을 배치 단위로 병렬 처리하는 함수 (url_edu.py 방식 적용)
async def process_chunks_parallel(
    all_chunks_info: list,
    content: str,
    subject: str,
    table_name: str,
    dbname: str,
    job_id: str,
    job_manager: AsyncJobManager,
    job_progress_manager: AsyncJobProgress,
    memo: str,
    chunk_progress: float,
    batch_size: int = 5,
    personal_info_filter: str = "N",  # 개인정보 필터링 옵션 추가
    content_type: str = "file",
):
    """청크들을 배치 단위로 병렬 처리"""
    
    processed_chunks = 0
    
    # 배치별로 청크 처리
    for i in range(0, len(all_chunks_info), batch_size):
        # 취소 확인
        status = await job_manager.get_job_status(job_id)
        if status == "cancel":
            logger.info(f"엑셀 배치 처리 중 작업 취소됨: job_id={job_id}")
            return processed_chunks

        batch_chunks = all_chunks_info[i:i + batch_size]
        batch_tasks = []

        # 배치 내 병렬 처리
        for chunk_info in batch_chunks:
            task = process_single_chunk_async(
                chunk_info=chunk_info,
                content=content,
                subject=subject,
                sheet_name=chunk_info["sheet_name"],
                table_name=table_name,
                dbname=dbname,
                memo=memo,
                job_id=job_id,
                job_manager=job_manager,
                personal_info_filter=personal_info_filter, # 개인정보 필터링 옵션 전달
            )
            batch_tasks.append(task)

        # 배치 실행
        try:
            batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
            
            # ✅ 배치 완료 후 진행률 업데이트
            batch_success_count = 0
            for result in batch_results:
                if isinstance(result, dict) and result.get("status") == "success":
                    batch_success_count += 1
            
            # ✅ 성공한 청크 수만큼 진행률 증가
            if batch_success_count > 0:
                processed_chunks += batch_success_count
                current_progress = await job_progress_manager.get_job_progress(job_id)
                progress_increment = chunk_progress * batch_success_count
                new_progress = round(min(current_progress + progress_increment, 99.99), 4)
                await job_progress_manager.set_job_progress(job_id, new_progress)
                

            
            logger.debug(f"엑셀 배치 처리 완료: {content}, 청크 {i+1}-{min(i+batch_size, len(all_chunks_info))}")
            
        except Exception as e:
            logger.error(f"엑셀 배치 처리 중 오류: {e}")
    
    return processed_chunks


def format_cell_value(cell):
    """셀의 서식에 따라 통화 기호나 숫자 서식을 적용"""
    value = cell.value
    if value is None:
        return ""
    if isinstance(cell, openpyxl.cell.cell.Cell):  # xlsx 파일의 경우
        if (
            "¥" in str(cell.number_format)
            or "₩" in str(cell.number_format)
            or "$" in str(cell.number_format)
            or "€" in str(cell.number_format)
        ):
            if isinstance(value, (int, float)):
                if "$" in str(cell.number_format):
                    return f"${value}"
                elif "₩" in str(cell.number_format):
                    return f"₩{value}"
                elif "¥" in str(cell.number_format):
                    return f"¥{value}"
                elif "€" in str(cell.number_format):
                    return f"€{value}"
    return str(value)


def load_excel_sheets_dict(file_path: str) -> Dict[str, List[List[Any]]]:
    """
    동기. .xlsx / .xls / .csv → { 시트명: [ [셀…], … ] }
    (기존 extract_excel_data 본문과 동일; 이벤트 루프 블로킹을 피하려면 asyncio.to_thread 로 감쌀 것)
    """
    file_extension = os.path.splitext(file_path)[1].lower()
    sheets_data: Dict[str, List[List[Any]]] = {}

    if file_extension == ".xlsx":
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            try:
                sheet = workbook[sheet_name]
                sheet_rows = []
                for row in sheet.iter_rows():
                    row_data = _normalize_tabular_row(
                        [format_cell_value(cell) for cell in row if cell.value is not None]
                    )
                    if row_data:
                        sheet_rows.append(row_data)
                if sheet_rows:
                    sheets_data[sheet_name] = sheet_rows
            except Exception as sheet_err:
                logger.warning("XLSX 시트 읽기 실패 | path=%s sheet=%s err=%s", file_path, sheet_name, sheet_err)

    elif file_extension == ".xls":
        if xlrd is None:
            raise ImportError("xlrd is required to process .xls files")
        workbook = xlrd.open_workbook(file_path)
        for sheet in workbook.sheets():
            try:
                sheet_rows = []
                for row_idx in range(sheet.nrows):
                    row = sheet.row(row_idx)
                    row_data = _normalize_tabular_row([cell.value for cell in row if cell.value not in (None, "")])
                    if row_data:
                        sheet_rows.append(row_data)
                if sheet_rows:
                    sheets_data[sheet.name] = sheet_rows
            except Exception as sheet_err:
                logger.warning("XLS 시트 읽기 실패 | path=%s sheet=%s err=%s", file_path, getattr(sheet, 'name', ''), sheet_err)

    elif file_extension == ".csv":
        with open(file_path, "rb") as f:
            result = chardet.detect(f.read())
            encoding = result.get("encoding") or "utf-8"

        csv_rows: List[List[Any]] = []
        df = None

        try:
            df = pd.read_csv(file_path, encoding=encoding, engine="python")
        except Exception as e1:
            logger.warning(f"CSV 기본 읽기 실패, 대안 방법 시도: {e1}")

            try:
                df = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    engine="python",
                    on_bad_lines="skip",
                    sep=None,
                )
                logger.info("CSV 읽기 성공 (잘못된 줄 건너뛰기 방식)")
            except Exception as e2:
                logger.warning(f"CSV 두 번째 방법 실패: {e2}")

                try:
                    with open(file_path, "r", encoding=encoding, errors="ignore") as csvfile:
                        for delimiter in [",", ";", "\t", "|"]:
                            try:
                                csvfile.seek(0)
                                reader = csv.reader(csvfile, delimiter=delimiter)
                                rows = list(reader)
                                first_row = _first_meaningful_row(rows)
                                if len(first_row) > 1:
                                    csv_rows = [
                                        _normalize_tabular_row(list(row or []))
                                        for row in rows
                                        if _normalize_tabular_row(list(row or []))
                                    ]
                                    logger.info(
                                        f"CSV 수동 읽기 성공 (구분자: '{delimiter}'), 총 {len(csv_rows)}행"
                                    )
                                    break
                            except Exception:
                                continue
                        else:
                            raise Exception("모든 구분자로 시도했지만 읽기 실패")
                except Exception as e3:
                    logger.error(f"CSV 수동 읽기도 실패: {e3}")
                    raise Exception(f"CSV 파일을 읽을 수 없습니다: {e1}") from e3

        if df is not None:
            for _, row in df.iterrows():
                row_data = _normalize_tabular_row([cell for cell in row if pd.notna(cell)])
                if row_data:
                    csv_rows.append(row_data)

        if csv_rows:
            sheets_data["csv"] = csv_rows

    else:
        raise ValueError(f"지원하지 않는 확장자: {file_extension}")

    return sheets_data


def extract_excel_to_plain_text(file_path: str) -> str:
    """파일 크롤 학습용: 시트별 행을 구분 가능한 평문 한 덩어리로 합친다."""
    try:
        sheets = load_excel_sheets_dict(file_path)
    except Exception as e:
        logger.warning("엑셀/CSV 평문 추출 실패 | path=%s err=%s", file_path, e)
        return ""
    if not sheets:
        return ""
    parts: List[str] = []
    for sheet_name, sheet_rows in sheets.items():
        parts.append(f"## {sheet_name}")
        for row in sheet_rows:
            parts.append(" | ".join(str(c).strip() for c in row))
        parts.append("")
    return "\n".join(parts).strip()


async def extract_excel_data(file_path: str):
    """
    엑셀 파일에서 각 시트별 데이터를 추출하여
    시트명을 key, 행 데이터 리스트를 value로 하는 딕셔너리로 반환
    """
    try:
        return await asyncio.to_thread(load_excel_sheets_dict, file_path)
    except Exception as e:
        logging.error(f"데이터 추출 중 오류 발생: {e}", exc_info=True)
        raise


def create_chunk_from_rows(sheet_name: str, header_rows: list, chunk_rows: list, chunk_idx: int) -> str:
    """
    행 데이터로부터 마크다운 테이블 형식의 청크를 생성합니다.
    """
    chunk_lines = [f"### Sheet: {sheet_name} ###"]
    
    # 헤더가 있는 경우 먼저 추가
    safe_header_rows = []
    for header_row in header_rows or []:
        normalized = _normalize_tabular_row(list(header_row or []))
        if normalized:
            safe_header_rows.append(normalized)

    if safe_header_rows:
        for header_row in safe_header_rows:
            chunk_lines.append("| " + " | ".join(map(str, header_row)) + " |")
        header_col_count = max((len(row) for row in safe_header_rows), default=0)
        if header_col_count > 0:
            chunk_lines.append("| " + " | ".join(["---"] * header_col_count) + " |")

    # 청크 데이터 추가
    for row_data in chunk_rows or []:
        normalized = _normalize_tabular_row(list(row_data or []))
        if normalized:
            chunk_lines.append("| " + " | ".join(map(str, normalized)) + " |")
    
    return "\n".join(chunk_lines)


async def process_xls(
    content: str,
    file_path: str,
    table_name: str,
    dbname: str,
    job_id: str,
    each_progress: float,
    job_manager: AsyncJobManager,
    job_progress_manager: AsyncJobProgress,
    subject: str = "",
    memo: str = "",
    personal_info_filter: str = "N",  # 개인정보 필터링 옵션 추가
    content_type: str = "file",
):
    """
    XLS/XLSX/CSV 파일을 30행 단위로 청킹하여 병렬 처리합니다.
    자체적으로 진행률을 계산하여 실시간 websocket 전송합니다.

    청크 상단 구조:
    ### Sheet: {시트명} ###
    {헤더 정보}
    {30행 데이터}
    """
    try:
        start_time = time.time()
        logger.info(f"[엑셀 병렬 처리 시작] 파일: {content}")
        
        # 🔍 1단계: 엑셀 데이터를 행 단위로 추출
        sheets_data = await await_file_text_extract(
            extract_excel_data(file_path),
            path=file_path,
            stage="excel",
            logger=logger,
        )
        
        ROWS_PER_CHUNK = 30  # 한 청크당 행 수
        total_chunks = 0
        all_chunks_info = []  # 모든 청크 정보를 하나의 리스트로 저장
        
        # 🔍 2단계: 각 시트별로 30행씩 청킹하여 총 청크 수 계산
        global_chunk_num = 0
        for sheet_idx, (sheet_name, sheet_rows) in enumerate(sheets_data.items(), start=1):
            if not sheet_rows:
                continue
                
            # 첫 번째 행을 헤더로 추정 (상위 2개 행 사용)
            header_rows = sheet_rows[:min(2, len(sheet_rows))]
            data_rows = sheet_rows[2:] if len(sheet_rows) > 2 else []
            
            # 30행씩 청킹
            for i in range(0, len(data_rows), ROWS_PER_CHUNK):
                chunk_data_rows = data_rows[i:i + ROWS_PER_CHUNK]
                global_chunk_num += 1
                chunk_info = {
                    "sheet_idx": sheet_idx,
                    "sheet_name": sheet_name,
                    "header_rows": header_rows,
                    "data_rows": chunk_data_rows,
                    "chunk_idx": i // ROWS_PER_CHUNK + 1,
                    "chunk_num": global_chunk_num,
                }
                all_chunks_info.append(chunk_info)
            
            # 데이터가 없어도 최소 1개 청크는 생성 (헤더만)
            if not data_rows:
                global_chunk_num += 1
                chunk_info = {
                    "sheet_idx": sheet_idx,
                    "sheet_name": sheet_name,
                    "header_rows": header_rows,
                    "data_rows": [],
                    "chunk_idx": 1,
                    "chunk_num": global_chunk_num,
                }
                all_chunks_info.append(chunk_info)
        
        total_chunks = len(all_chunks_info)

        if total_chunks == 0:
            logger.warning(f"처리할 청크가 없습니다: {content}, job_id:{job_id}")
            return {"chunks": 0}

        # 🎯 3단계: 자체 진행률 계산 (100 / 실제청크수)
        chunk_progress = round(100.0 / total_chunks, 4)
        logger.info(f"[엑셀 병렬 청킹] {content}: 총 {sum(len(rows) for rows in sheets_data.values())}행 → {total_chunks}개 청크 (30행/청크), 청크당 진행률: {chunk_progress}%")
        
        # 🚀 4단계: 병렬 청크 처리 실행
        processed_chunks = await process_chunks_parallel(
            all_chunks_info=all_chunks_info,
            content=content,
            subject=subject or os.path.basename(file_path or "") or content,
            table_name=table_name,
            dbname=dbname,
            job_id=job_id,
            job_manager=job_manager,
            job_progress_manager=job_progress_manager,
            memo=memo,
            chunk_progress=chunk_progress,
            batch_size=FILE_EMBEDDING_BATCH_SIZE,
            personal_info_filter=personal_info_filter, # 개인정보 필터링 옵션 전달
            content_type=content_type,
        )

        processing_time = round(time.time() - start_time, 2)
        logger.info(f"[엑셀 병렬 처리 완료] {content}: {total_chunks}개 청크 중 {processed_chunks}개 처리 완료 (30행/청크 기준), 처리 시간: {processing_time}초")
        
        return {
            "status": "success",
            "chunks": processed_chunks,
            "chunk_count": [processed_chunks],
            "use_source": [content]
        }

    except Exception as e:
        logging.error(f"엑셀 파일 처리 중 오류 발생: {e}", exc_info=True)

        raise RuntimeError(f"엑셀 파일 처리 중 오류 발생: {e}")


async def calculate_excel_chunks(file_path: str) -> int:
    """Excel 파일의 청크 수 예상치 (실제 계산은 xls_edu.py에서 담당)"""
    try:
        from backend.config import Config
        
        # 🎯 단순 예상치만 반환 (실제 계산은 process_excel에서 담당)
        file_size = os.path.getsize(file_path)
        estimated_chunks = max(1, file_size // (Config.BASIC_CHUNK_SIZE * 50))  # 대략적 추정
        logger.info(f"엑셀 예상 청크: {file_path}, 예상: {estimated_chunks}개 (실제는 process_excel에서 계산)")
        return estimated_chunks
    except Exception as e:
        logger.error(f"Excel 청크 수 추정 실패: {file_path}, 오류: {e}")
        return 10  # 기본값
